"""Build .xlsx fixtures: protected, unprotected, and multi-sheet.

Sheet and workbook protection are advisory flags in the OOXML, not encryption:
the cell values are readable either way. These fixtures exercise that case.

    python tests/make_xlsx_fixture.py out.xlsx [--open|--multi]
"""

import sys
from pathlib import Path

from openpyxl import Workbook


def _fill(ws, title: str) -> None:
    ws["A1"] = title
    ws["A2"] = "Second row of the sheet body."
    ws["B1"] = 42
    ws["B2"] = 3.14


def build_protected(path: Path, password: str = "secret") -> Path:
    """Sheet protection plus workbook structure protection."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Locked"
    _fill(ws, "This sheet is protected.")

    ws.protection.set_password(password)
    ws.protection.sheet = True
    ws.protection.enable()

    # Structure protection stops sheets being added, renamed or deleted.
    wb.security.workbookPassword = password
    wb.security.lockStructure = True

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def build_open(path: Path) -> Path:
    """No protection at all."""
    wb = Workbook()
    _fill(wb.active, "This sheet has no protection.")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def build_multi(path: Path, password: str = "secret") -> Path:
    """Several sheets, only some protected."""
    wb = Workbook()

    first = wb.active
    first.title = "Locked1"
    _fill(first, "Protected sheet one.")
    first.protection.set_password(password)
    first.protection.sheet = True
    first.protection.enable()

    second = wb.create_sheet("Open")
    _fill(second, "This one is not protected.")

    third = wb.create_sheet("Locked2")
    _fill(third, "Protected sheet two.")
    third.protection.set_password(password)
    third.protection.sheet = True
    third.protection.enable()

    wb.security.workbookPassword = password
    wb.security.lockStructure = True

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


if __name__ == "__main__":
    out = Path(sys.argv[1] if len(sys.argv) > 1 else "tests/fixtures/protected.xlsx")
    if "--open" in sys.argv:
        build_open(out)
    elif "--multi" in sys.argv:
        build_multi(out)
    else:
        build_protected(out)
    print(f"wrote {out} ({out.stat().st_size} bytes)")
