"""Tests for .xlsx protection removal."""

import re
import sys
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from backend import dispatch  # noqa: E402
from backend.excel_cracker import ExcelCracker, inspect_xlsx  # noqa: E402
from tests.make_xlsx_fixture import build_multi, build_open, build_protected  # noqa: E402


@pytest.fixture
def protected(tmp_path):
    return build_protected(tmp_path / "protected.xlsx")


@pytest.fixture
def unprotected(tmp_path):
    return build_open(tmp_path / "open.xlsx")


@pytest.fixture
def multi(tmp_path):
    return build_multi(tmp_path / "multi.xlsx")


def count_protection(path) -> tuple:
    """(sheetProtection count, workbookProtection count) in a workbook."""
    sheets = book = 0
    with zipfile.ZipFile(path) as zf:
        for name in zf.namelist():
            lowered = name.lower()
            if lowered.startswith(("xl/worksheets/", "xl/chartsheets/")) and lowered.endswith(".xml"):
                sheets += len(re.findall(rb"<sheetProtection", zf.read(name)))
            elif lowered == "xl/workbook.xml":
                book += len(re.findall(rb"<workbookProtection", zf.read(name)))
    return sheets, book


# ---------------------------------------------------------------- core


def test_removes_sheet_and_workbook_protection(protected, tmp_path):
    assert count_protection(protected) == (1, 1)

    result = ExcelCracker(str(protected), str(tmp_path / "out")).unlock()
    assert result.status == "success", result.error
    assert result.file_format == "xlsx"
    assert result.protections_found == 2
    assert count_protection(result.output_path) == (0, 0)


def test_original_is_untouched(protected, tmp_path):
    before = protected.read_bytes()
    ExcelCracker(str(protected), str(tmp_path / "out")).unlock()
    assert protected.exists()
    assert protected.read_bytes() == before


def test_cell_data_survives(protected, tmp_path):
    """Unlocking must not disturb the workbook's contents."""
    result = ExcelCracker(str(protected), str(tmp_path / "out")).unlock()
    wb = load_workbook(result.output_path)
    ws = wb["Locked"]
    assert ws["A1"].value == "This sheet is protected."
    assert ws["A2"].value == "Second row of the sheet body."
    assert ws["B1"].value == 42
    assert ws["B2"].value == 3.14


def test_unlocked_workbook_reports_no_protection(protected, tmp_path):
    result = ExcelCracker(str(protected), str(tmp_path / "out")).unlock()
    wb = load_workbook(result.output_path)
    assert wb["Locked"].protection.sheet is False
    assert not wb.security or not wb.security.lockStructure


def test_every_protected_sheet_is_unlocked(multi, tmp_path):
    """A workbook with several sheets must have all of them cleared."""
    assert count_protection(multi)[0] == 2

    result = ExcelCracker(str(multi), str(tmp_path / "out")).unlock()
    assert result.status == "success", result.error
    assert count_protection(result.output_path) == (0, 0)

    wb = load_workbook(result.output_path)
    assert set(wb.sheetnames) == {"Locked1", "Open", "Locked2"}
    for name in wb.sheetnames:
        assert wb[name].protection.sheet is False


def test_entry_order_is_preserved(protected, tmp_path):
    """Reshuffling entries is what triggers Excel's repair prompt.

    The rule is to preserve whatever order the source used -- not to impose
    one. Word puts [Content_Types].xml first; openpyxl puts it last, and Excel
    accepts both.
    """
    result = ExcelCracker(str(protected), str(tmp_path / "out")).unlock()
    with zipfile.ZipFile(protected) as src, zipfile.ZipFile(result.output_path) as out:
        assert src.namelist() == out.namelist()
        assert "[Content_Types].xml" in out.namelist()
        assert out.testzip() is None


def test_unrelated_parts_are_byte_identical(protected, tmp_path):
    """Only the sheets and workbook.xml should differ."""
    result = ExcelCracker(str(protected), str(tmp_path / "out")).unlock()
    with zipfile.ZipFile(protected) as src, zipfile.ZipFile(result.output_path) as out:
        for name in src.namelist():
            lowered = name.lower()
            if lowered == "xl/workbook.xml" or lowered.startswith("xl/worksheets/"):
                continue
            assert src.read(name) == out.read(name), f"{name} changed unexpectedly"


def test_unprotected_workbook_still_succeeds(unprotected, tmp_path):
    result = ExcelCracker(str(unprotected), str(tmp_path / "out")).unlock()
    assert result.status == "success"
    assert count_protection(result.output_path)[0] == 0


def test_rejects_password_protected_workbook(tmp_path):
    """A password-to-open workbook is an encrypted OLE file, not an .xlsx ZIP."""
    p = tmp_path / "encrypted.xlsx"
    p.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 512)
    result = ExcelCracker(str(p)).unlock()
    assert result.status == "failed"
    assert "password-protected" in result.error


def test_rejects_garbage(tmp_path):
    p = tmp_path / "fake.xlsx"
    p.write_bytes(b"this is not a zip")
    result = ExcelCracker(str(p)).unlock()
    assert result.status == "failed"


def test_rejects_empty_file(tmp_path):
    p = tmp_path / "empty.xlsx"
    p.touch()
    assert "empty" in ExcelCracker(str(p)).unlock().error.lower()


def test_rejects_missing_file(tmp_path):
    result = ExcelCracker(str(tmp_path / "ghost.xlsx")).unlock()
    assert "not found" in result.error.lower()


def test_rejects_zip_without_a_workbook(tmp_path):
    p = tmp_path / "notexcel.xlsx"
    with zipfile.ZipFile(p, "w") as zf:
        zf.writestr("hello.txt", "not a workbook")
    result = ExcelCracker(str(p)).unlock()
    assert result.status == "failed"
    assert "workbook.xml" in result.error


def test_logs_cover_all_seven_steps(protected, tmp_path):
    result = ExcelCracker(str(protected), str(tmp_path / "out")).unlock()
    joined = "\n".join(result.logs)
    for n in range(1, 8):
        assert f"[{n}/7]" in joined


def test_no_temp_directory_leaks(protected, tmp_path):
    cracker = ExcelCracker(str(protected), str(tmp_path / "out"))
    cracker.unlock()
    assert cracker._temp_dir is not None
    assert not cracker._temp_dir.exists()


def test_second_run_does_not_clobber(protected, tmp_path):
    out = tmp_path / "out"
    first = ExcelCracker(str(protected), str(out)).unlock()
    second = ExcelCracker(str(protected), str(out)).unlock()
    assert first.output_path != second.output_path
    assert Path(first.output_path).exists() and Path(second.output_path).exists()


# ------------------------------------------------------------- inspect


def test_inspect_reports_protection(protected, unprotected):
    info = inspect_xlsx(str(protected))
    assert info["format"] == "xlsx"
    assert info["protected"] is True
    assert info["can_unlock"] is True
    assert info["sheets"]

    clean = inspect_xlsx(str(unprotected))
    assert clean["sheets"] == []


# ------------------------------------------------------------ dispatch


def test_dispatch_routes_xlsx(protected, tmp_path):
    result = dispatch.unlock(str(protected), str(tmp_path / "out"))
    assert result.status == "success"
    assert result.file_format == "xlsx"
    assert result.method == "ooxml"


def test_dispatch_detects_xlsx():
    assert dispatch.detect_format("book.xlsx") == "xlsx"
    assert dispatch.detect_format("book.XLSX") == "xlsx"
    assert dispatch.format_label("book.xlsx") == "XLSX"
    assert dispatch.STEP_COUNT["xlsx"] == 7


def test_dispatch_inspect_xlsx(protected):
    assert dispatch.inspect(str(protected))["format"] == "xlsx"
